"""load_county_rucc.py — populate counties.rucc_code from USDA ERS
Rural-Urban Continuum Codes (2023 vintage, ~3,143 rows).

Source: https://www.ers.usda.gov/data-products/rural-urban-continuum-codes/
File: Ruralurbancontinuumcodes2023.xlsx — sheet "Rural-urban Continuum Code 2023"

Idempotent: re-running just refreshes the columns.

Usage:
  python scripts/one_off/load_county_rucc.py
"""
from __future__ import annotations
import argparse, os, subprocess, sys, time, urllib.parse
from pathlib import Path
import psycopg2
from dotenv import load_dotenv

ROOT = Path(__file__).resolve().parent.parent.parent
load_dotenv(ROOT / 'scripts' / 'pipeline' / '.env')
POOLER = (ROOT / 'supabase' / '.temp' / 'pooler-url').read_text().strip()
_p = urllib.parse.urlparse(POOLER)
PG = dict(host=_p.hostname, port=_p.port or 5432, user=_p.username,
          password=os.environ['SUPABASE_DB_PASSWORD'],
          dbname=(_p.path or '/postgres').lstrip('/'), sslmode='require')

CACHE_DIR = ROOT / 'supabase' / '.temp' / 'rucc'
RUCC_URL = ('https://www.ers.usda.gov/sites/default/files/_laserfiche/'
            'DataFiles/53251/Ruralurbancontinuumcodes2023.xlsx')
RUCC_VINTAGE = 2023

RUCC_LABELS = {
    1: 'Metro: county in metro area of 1M+',
    2: 'Metro: county in metro area of 250K-1M',
    3: 'Metro: county in metro area of <250K',
    4: 'Nonmetro: urban pop 20K+, adjacent to metro',
    5: 'Nonmetro: urban pop 20K+, not adjacent',
    6: 'Nonmetro: urban pop 2.5K-20K, adjacent',
    7: 'Nonmetro: urban pop 2.5K-20K, not adjacent',
    8: 'Nonmetro: <2.5K or rural, adjacent',
    9: 'Nonmetro: <2.5K or rural, not adjacent',
}


def log(m: str): print(f"[{time.strftime('%H:%M:%S')}] {m}", flush=True)


def download() -> Path | None:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    out = CACHE_DIR / 'Ruralurbancontinuumcodes2023.xlsx'
    if out.exists() and out.stat().st_size > 100_000:
        log(f'    using cached {out}')
        return out
    log(f'    downloading {RUCC_URL}')
    rc = subprocess.run(
        ['curl', '--ssl-no-revoke', '-sSL', '-o', str(out), RUCC_URL],
        capture_output=True
    )
    if rc.returncode != 0 or not out.exists() or out.stat().st_size < 100_000:
        log(f'    download failed (rc={rc.returncode}, size={out.stat().st_size if out.exists() else 0})')
        return None
    return out


def parse_xlsx(path: Path) -> list[tuple[str, int]]:
    """Return list of (geoid_5char, rucc_code)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        log('    pip install openpyxl required'); sys.exit(2)
    wb = load_workbook(path, data_only=True, read_only=True)
    # USDA may name the sheet variously across vintages; try common names.
    candidates = [n for n in wb.sheetnames
                  if 'rural' in n.lower() or 'rucc' in n.lower() or '2023' in n]
    sheet_name = candidates[0] if candidates else wb.sheetnames[0]
    ws = wb[sheet_name]
    log(f'    parsing sheet: {sheet_name}')

    rows = ws.iter_rows(values_only=True)
    header = None
    out = []
    for row in rows:
        if header is None:
            # Header row contains both 'FIPS' and 'RUCC' column names
            if any(c and 'FIPS' in str(c).upper() for c in row) \
               and any(c and 'RUCC' in str(c).upper() for c in row):
                header = list(row)
                # Find column indices
                fips_idx = next(i for i, c in enumerate(header)
                                if c and 'FIPS' in str(c).upper())
                rucc_idx = next(i for i, c in enumerate(header)
                                if c and 'RUCC' in str(c).upper() and '2023' in str(c))
            continue
        if not row or row[fips_idx] is None or row[rucc_idx] is None:
            continue
        try:
            geoid = str(row[fips_idx]).strip().zfill(5)
            rucc  = int(row[rucc_idx])
        except (ValueError, TypeError):
            continue
        if rucc < 1 or rucc > 9:
            continue
        out.append((geoid, rucc))
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    path = download()
    if path is None:
        sys.exit(1)
    log(f'Parsing {path.name}')
    rows = parse_xlsx(path)
    log(f'  {len(rows)} (county_fips, rucc) tuples parsed')
    if args.dry_run:
        for r in rows[:5]: print(' ', r)
        return

    log('Updating counties.rucc_*')
    with psycopg2.connect(**PG) as c:
        c.autocommit = True
        with c.cursor() as cur:
            n_updated = 0
            for geoid, rucc in rows:
                cur.execute("""
                    update public.counties
                       set rucc_code = %s,
                           rucc_label = %s,
                           rucc_vintage = %s
                     where geoid = %s
                """, (rucc, RUCC_LABELS[rucc], RUCC_VINTAGE, geoid))
                if cur.rowcount > 0:
                    n_updated += 1
            cur.execute("select count(*), count(rucc_code) from public.counties")
            total, with_rucc = cur.fetchone()
    log(f'Done. updated={n_updated} of {total} counties; total_with_rucc={with_rucc}')


if __name__ == '__main__':
    main()
